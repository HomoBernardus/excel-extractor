#!/usr/bin/env python3
"""
分箱单生成程序
数据源: 总单.xlsx (GPL sheet) + 附件备件清单.xls
模板: 分箱单.xlsx
输出: 分箱单_生成结果.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.pagebreak import Break
from copy import copy
import xlrd
import re
import os

# ── Paths ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, '分箱单.xlsx')
ZONGDAN_PATH  = os.path.join(BASE_DIR, '总单.xlsx')
SPARE_PARTS_PATH = os.path.join(BASE_DIR, '附件备件清单.xls')
OUTPUT_PATH   = os.path.join(BASE_DIR, '分箱单_生成结果.xlsx')

TEMPLATE_ROWS = 18
TEMPLATE_COLS = 6

# ── Merged cell definitions from template (offsets relative to page start) ──
# (row_offset from page_start, col1, row2_offset, col2)
MERGED_RANGES = [
    (0, 1, 0, 6),   # A1:F1
    (1, 1, 1, 6),   # A2:F2
    (2, 1, 2, 2),   # A3:B3
    (2, 4, 2, 5),   # D3:E3
    (3, 1, 3, 2),   # A4:B4
    (3, 4, 3, 5),   # D4:E4
    (4, 1, 4, 2),   # A5:B5
    (4, 4, 4, 6),   # D5:F5
    (5, 1, 5, 2),   # A6:B6
]


def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def read_gpl_data(zongdan_path):
    """从 总单.xlsx GPL sheet 读取数据."""
    wb = openpyxl.load_workbook(zongdan_path)
    ws = wb['GPL']

    # 找 MNS… 边界行
    mns_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value and 'MNS Low Voltage Switchgear' in str(cell.value):
                mns_row = cell.row
                break
        if mns_row:
            break

    if mns_row is None:
        raise ValueError("未找到 MNS Low Voltage Switchgear 边界行")

    # C17 = row 17, col 3 → Customer PO No.
    customer_po = ws.cell(row=17, column=3).value

    # 数据从 MNS 行下一行开始
    data_rows = []
    for r in range(mns_row + 1, ws.max_row + 1):
        case_no = ws.cell(row=r, column=2).value       # B
        if case_no is None:
            break
        data_rows.append({
            'case_no':    ws.cell(row=r, column=2).value,  # B
            'description': ws.cell(row=r, column=3).value,  # C
            'qty':        ws.cell(row=r, column=4).value,  # D
            'size':       ws.cell(row=r, column=5).value,  # E
            'nw':         ws.cell(row=r, column=6).value,  # F
            'gw':         ws.cell(row=r, column=7).value,  # G
            'system_name': ws.cell(row=r, column=8).value,  # H
            'contract_no': ws.cell(row=r, column=9).value,  # I
        })

    wb.close()
    return customer_po, data_rows


def parse_contract_pattern(contract_no):
    """解析合同号, 返回 (base, variant_number).
    如 PM009690-01-01FJ1 → ('PM009690-01-01FJ', '1')"""
    contract_str = str(contract_no).strip()
    m = re.match(r'^(.+[A-Za-z])(\d+\.*)$', contract_str)
    if m:
        return m.group(1), m.group(2).rstrip('.')
    return contract_str, ''


def find_matching_sheets(spare_sheet_names, contract_no):
    """根据合同号匹配附件备件清单中的 sheet."""
    base, variant_num = parse_contract_pattern(contract_no)
    if not variant_num:
        return []

    matching = []
    for sname in spare_sheet_names:
        sb, sv = parse_contract_pattern(sname)
        if sb == base and sv == variant_num:
            matching.append(sname)

    def sort_key(name):
        _, sv = parse_contract_pattern(name)
        return re.match(r'^(.+[A-Za-z])(\d+\.*)$', name).group(2)

    matching.sort(key=sort_key)
    return matching


def read_spare_parts(spare_wb, sheet_name):
    """读取单个备件 sheet 的数据行."""
    sheet = spare_wb.sheet_by_name(sheet_name)
    items = []
    for r in range(3, sheet.nrows):  # 数据从第4行(index 3)开始
        desc = sheet.cell_value(r, 1)
        typ  = sheet.cell_value(r, 2)
        qty  = sheet.cell_value(r, 3)
        unit = sheet.cell_value(r, 4)
        remarks = sheet.cell_value(r, 6)

        # 只保留有数值型数量的行 (元数据行无数量)
        try:
            qty_num = float(qty)
            if qty_num <= 0:
                continue
        except (ValueError, TypeError):
            continue

        items.append({
            'description': desc,
            'type': typ,
            'qty': int(qty_num) if qty_num == int(qty_num) else qty_num,
            'unit': unit,
            'remarks': remarks,
        })
    return items


def add_merged_cells(ws, page_start_row):
    """为当前页添加合并单元格."""
    for r1_off, c1, r2_off, c2 in MERGED_RANGES:
        ws.merge_cells(
            start_row=page_start_row + r1_off,
            start_column=c1,
            end_row=page_start_row + r2_off,
            end_column=c2
        )


def copy_template_page(template_ws, out_ws, page_start_row):
    """复制模板的 18 行到输出 sheet 的指定位置."""
    for offset in range(TEMPLATE_ROWS):
        out_row = page_start_row + offset
        tpl_row = offset + 1  # 1-indexed in template

        # 复制行高
        tpl_height = template_ws.row_dimensions[tpl_row].height
        if tpl_height:
            out_ws.row_dimensions[out_row].height = tpl_height

        for col in range(1, TEMPLATE_COLS + 1):
            src = template_ws.cell(row=tpl_row, column=col)
            dst = out_ws.cell(row=out_row, column=col)
            copy_cell_style(src, dst)
            dst.value = src.value

    # 重建合并单元格
    add_merged_cells(out_ws, page_start_row)

    # 统一 E 列(Unit)数据行垂直居中 (模板 16-18 行不一致)
    for offset in range(9, TEMPLATE_ROWS):  # Row 10 起
        cell = out_ws.cell(row=page_start_row + offset, column=5)
        current = cell.alignment
        cell.alignment = Alignment(
            horizontal=current.horizontal or 'center',
            vertical='center',
            wrap_text=current.wrap_text,
        )


def fill_page_data(out_ws, page_start_row, customer_po, row_data, is_acc=False):
    """填入单页数据字段."""
    # C3: Customer PO
    out_ws.cell(row=page_start_row + 2, column=3).value = customer_po
    # F3: Case No.
    case_no = row_data['case_no']
    out_ws.cell(row=page_start_row + 2, column=6).value = int(case_no) if isinstance(case_no, float) and case_no == int(case_no) else case_no
    # C4: Size
    sz = row_data['size']
    out_ws.cell(row=page_start_row + 3, column=3).value = str(sz) if sz is not None else ''
    # C5: Net Weight
    nw = row_data['nw']
    out_ws.cell(row=page_start_row + 4, column=3).value = int(nw) if isinstance(nw, float) and nw == int(nw) else nw
    # C6: Gross Weight
    gw = row_data['gw']
    out_ws.cell(row=page_start_row + 5, column=3).value = int(gw) if isinstance(gw, float) and gw == int(gw) else gw
    # F4: Project/Contract No.
    out_ws.cell(row=page_start_row + 3, column=6).value = str(row_data['contract_no']) if row_data['contract_no'] is not None else ''

    # 规格/数量/系统号: Accessories 填在 Row 9, 普通填在 Row 10
    data_offset = 8 if is_acc else 9

    desc = row_data['description']
    out_ws.cell(row=page_start_row + data_offset, column=3).value = str(desc) if desc is not None else ''
    qty = row_data['qty']
    if qty is not None:
        out_ws.cell(row=page_start_row + data_offset, column=4).value = int(qty) if isinstance(qty, float) and qty == int(qty) else qty
    sys_name = row_data['system_name']
    out_ws.cell(row=page_start_row + data_offset, column=6).value = str(sys_name) if sys_name is not None else ''


def delete_mns_row(out_ws, page_start_row):
    """删除 Row 9 (MNS 行), 将 Row 10-18 上移一格, Row 18 清空."""
    for offset in range(9, 18):
        src_row = page_start_row + offset      # 10 ~ 18
        dst_row = page_start_row + offset - 1  # 9 ~ 17
        out_ws.row_dimensions[dst_row].height = out_ws.row_dimensions[src_row].height
        for col in range(1, TEMPLATE_COLS + 1):
            src = out_ws.cell(row=src_row, column=col)
            dst = out_ws.cell(row=dst_row, column=col)
            copy_cell_style(src, dst)
            dst.value = src.value

    # 清空 Row 18
    last_row = page_start_row + 17
    out_ws.row_dimensions[last_row].height = 18.6
    for col in range(1, TEMPLATE_COLS + 1):
        cell = out_ws.cell(row=last_row, column=col)
        cell.value = None
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )


def fill_spare_parts(out_ws, page_start_row, spare_items, template_ws, start_offset=10):
    """填入备件明细.

    Args:
        start_offset: spare parts 起始行的 page offset
                      (普通页 10=Row11, Accessories 页 9=Row10)
    """
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    style_ref_row = page_start_row + start_offset - 1  # 数据描述行

    for i, item in enumerate(spare_items):
        data_row = page_start_row + start_offset + i

        # 如果超过模板行数, 需要新建行
        rows_in_template = TEMPLATE_ROWS - start_offset
        if i >= rows_in_template:
            for col_idx in range(1, TEMPLATE_COLS + 1):
                dst = out_ws.cell(row=data_row, column=col_idx)
                src_style = out_ws.cell(row=style_ref_row, column=col_idx)
                copy_cell_style(src_style, dst)
                dst.value = None
            out_ws.row_dimensions[data_row].height = out_ws.row_dimensions[style_ref_row].height

        # A: 序号
        out_ws.cell(row=data_row, column=1).value = i + 1
        # B: Description (物料号在备用清单 col 1)
        val = item['description']
        if val and str(val).strip():
            out_ws.cell(row=data_row, column=2).value = str(val)
        # C: Type (型号在备用清单 col 2)
        val = item['type']
        if val and str(val).strip():
            out_ws.cell(row=data_row, column=3).value = str(val)
        # D: Qty (数量)
        qty = item['qty']
        if qty != '' and qty is not None:
            try:
                out_ws.cell(row=data_row, column=4).value = int(float(qty))
            except (ValueError, TypeError):
                out_ws.cell(row=data_row, column=4).value = qty
        # E: Unit
        val = item['unit']
        if val and str(val).strip():
            out_ws.cell(row=data_row, column=5).value = str(val)
        # F: Remarks
        val = item['remarks']
        if val and str(val).strip():
            out_ws.cell(row=data_row, column=6).value = str(val)

        # 有内容的单元格补边框
        for col_idx in range(1, TEMPLATE_COLS + 1):
            cell = out_ws.cell(row=data_row, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                cell.border = thin_border


def adjust_row_heights_for_wrap(out_ws, page_start_row, page_rows):
    """C 列(品名规格)或 F 列(备注)文本较长时, 行高设为原来的 2 倍."""
    for r_off in range(9, page_rows):  # 从 Row 10 (数据区) 开始
        row = page_start_row + r_off
        # 检查 C 列(3) 和 F 列(6)
        c_val = out_ws.cell(row=row, column=3).value
        f_val = out_ws.cell(row=row, column=6).value
        c_len = len(str(c_val)) if c_val else 0
        f_len = len(str(f_val)) if f_val else 0

        if c_len > 40 or f_len > 40:
            current_height = out_ws.row_dimensions[row].height
            if current_height:
                out_ws.row_dimensions[row].height = current_height * 2


def add_row_borders_for_content(out_ws, page_start_row, page_rows):
    """为有文字的区域补全边框."""
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for r_off in range(page_rows):
        row = page_start_row + r_off
        for col in range(1, TEMPLATE_COLS + 1):
            cell = out_ws.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip():
                # 确保边框存在
                if cell.border.left.style is None:
                    cell.border = thin_border


def generate(template_path=None, zongdan_path=None, spare_path=None,
             output_path=None, progress_callback=None):
    """生成分箱单.

    Args:
        progress_callback: 可选, 签名为 callback(message, current, total)
    """
    tpl_path = template_path or TEMPLATE_PATH
    zd_path  = zongdan_path or ZONGDAN_PATH
    sp_path  = spare_path or SPARE_PARTS_PATH
    out_path = output_path or OUTPUT_PATH

    def log(msg, current=0, total=0):
        if progress_callback:
            progress_callback(msg, current, total)
        else:
            print(msg)

    log("读取模板...")
    tpl_wb = openpyxl.load_workbook(tpl_path)
    tpl_ws = tpl_wb.active

    log("读取总单数据...")
    customer_po, data_rows = read_gpl_data(zd_path)
    log(f"  Customer PO: {customer_po}")
    log(f"  数据行数: {len(data_rows)}")
    total_pages = len(data_rows)

    log("读取附件备件清单...")
    spare_wb = xlrd.open_workbook(sp_path)
    spare_sheet_names = spare_wb.sheet_names()

    log("生成分箱单...")
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = '分箱单'

    for col in range(1, TEMPLATE_COLS + 1):
        col_letter = get_column_letter(col)
        out_ws.column_dimensions[col_letter].width = \
            tpl_ws.column_dimensions[col_letter].width

    current_row = 1
    page_breaks = []

    for idx, row_data in enumerate(data_rows):
        page_start = current_row
        is_acc = str(row_data['description']).strip().lower() == 'accessories'

        spare_items = []
        if is_acc:
            contract_no = str(row_data['contract_no'])
            matching = find_matching_sheets(spare_sheet_names, contract_no)
            for sname in matching:
                spare_items.extend(read_spare_parts(spare_wb, sname))
            page_rows = max(TEMPLATE_ROWS, 9 + len(spare_items))
            log(f"  页 {idx+1}/{total_pages}: Case {row_data['case_no']} - Accessories ({len(spare_items)} 条)",
                idx + 1, total_pages)
        else:
            page_rows = TEMPLATE_ROWS
            log(f"  页 {idx+1}/{total_pages}: Case {row_data['case_no']} - {row_data['description']}",
                idx + 1, total_pages)

        copy_template_page(tpl_ws, out_ws, page_start)

        if is_acc:
            delete_mns_row(out_ws, page_start)

        fill_page_data(out_ws, page_start, customer_po, row_data, is_acc=is_acc)

        if is_acc and spare_items:
            fill_spare_parts(out_ws, page_start, spare_items, tpl_ws, start_offset=9)

        adjust_row_heights_for_wrap(out_ws, page_start, page_rows)
        add_row_borders_for_content(out_ws, page_start, page_rows)

        if idx < total_pages - 1:
            page_breaks.append(current_row + page_rows - 1)

        current_row += page_rows

    for pb in page_breaks:
        out_ws.row_breaks.append(Break(id=pb))

    out_ws.sheet_properties.pageSetUpPr = \
        openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    out_wb.save(out_path)
    log(f"\n完成! 输出: {out_path}")
    log(f"  总页数: {total_pages}, 总行数: {current_row - 1}")

    tpl_wb.close()
    spare_wb.release_resources()
    out_wb.close()


if __name__ == '__main__':
    generate()
